# -*- coding: utf-8 -*-
"""
Document Automation System (DAS) - Flask Backend V3.0
Nhập liệu Tiếng Việt → Hệ thống quản lý và xuất hồ sơ thông minh.
"""
import os, uuid, re, unicodedata, json, base64, traceback, io, zipfile, requests
from datetime import date, datetime, timedelta, timezone
from flask import Flask, request, jsonify, send_file, render_template, Response, make_response
from flask_cors import CORS
from jinja2 import Template
from deep_translator import GoogleTranslator
from dotenv import load_dotenv
from urllib.parse import quote
from unicodedata import normalize
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, text, inspect
from flask_basicauth import BasicAuth
from PIL import Image
from vietnamese_names_dict import get_vietnamese_name_in_chinese
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import google.generativeai as genai

load_dotenv()

# Configure Gemini
gemini_api_key = os.environ.get('GEMINI_API_KEY')
if gemini_api_key:
    genai.configure(api_key=gemini_api_key)

app = Flask(__name__, static_folder='static', static_url_path='')
app.debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
CORS(app, resources={r"/*": {"origins": ["https://cv.fct.vn", "http://127.0.0.1:5000", "http://localhost:5000"]}})
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20MB limit

app.config['BASIC_AUTH_USERNAME'] = os.environ.get('ADMIN_USERNAME', 'fctvt')
app.config['BASIC_AUTH_PASSWORD'] = os.environ.get('ADMIN_PASSWORD', '1503')
app.config['BASIC_AUTH_FORCE_PROMPT'] = True
basic_auth = BasicAuth(app)

def auth_required(f):
    if os.environ.get('RENDER'):
        return basic_auth.required(f)
    return f

# --- DATABASE CONFIG ---
db_url = os.environ.get('DATABASE_URL')
if db_url and db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
if not db_url:
    db_url = 'sqlite:///' + os.path.join(os.path.dirname(os.path.abspath(__file__)), 'history.db')

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Neon PostgreSQL: tránh lỗi SSL connection closed unexpectedly
# pool_pre_ping: kiểm tra connection còn sống trước mỗi query
# pool_recycle: tái tạo connection sau 300s để tránh idle timeout
# pool_size / max_overflow: giới hạn số connection
if db_url and db_url.startswith("postgresql"):
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,
        'pool_recycle': 300,
        'pool_size': 5,
        'max_overflow': 2,
        'connect_args': {
            'sslmode': 'require',
            'connect_timeout': 10,
            'keepalives': 1,
            'keepalives_idle': 30,
            'keepalives_interval': 10,
            'keepalives_count': 5,
        }
    }

db = SQLAlchemy(app)

class FormHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ma_so = db.Column(db.String(50))
    ho_ten = db.Column(db.String(100))
    ten_file = db.Column(db.String(255))
    data_json = db.Column(db.Text)
    ngay_tao = db.Column(db.DateTime, default=datetime.utcnow)
    is_selected = db.Column(db.Boolean, default=False)  # Trúng tuyển
    is_deleted = db.Column(db.Boolean, default=False)   # Xóa mềm
    don_hang = db.Column(db.String(255), default='')    # Đơn hàng ứng tuyển (mã đơn)
    nguoi_phu_trach = db.Column(db.String(100), default='') # Tên Nhân viên - Đối tác

class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)

def normalize_npt(f48_raw):
    f48 = str(f48_raw or '').strip()
    if not f48: return ""
    if '-' in f48:
        parts = f48.split('-', 1)
        emp = parts[0].strip()
        partner = parts[1].strip().title() if parts[1].strip() else ""
    else:
        parts = [p.strip() for p in f48.split()]
        if len(parts) >= 2:
            emp = " ".join(parts[:-1])
            partner = parts[-1].title()
        else:
            emp = f48
            partner = ""
            
    if emp.lower() == 'javiko' and not partner:
        emp, partner = 'Vũ', 'Javiko'
    elif partner.lower() == 'javiko' and not emp:
        emp, partner = 'Vũ', 'Javiko'
    elif (emp + partner).lower() == 'javiko':
        emp, partner = 'Vũ', 'Javiko'
        
    if emp.lower() in ('vũ', 'vu'):
        emp = 'AT'
        
    if emp and partner:
        return f"{emp} - {partner}"
    elif emp:
        return emp
    elif partner:
        return partner
    return ""

with app.app_context():
    try:
        db.create_all()
        # Auto migration for is_deleted, don_hang, nguoi_phu_trach
        try:
            inspector = inspect(db.engine)
            # Lấy danh sách cột của bảng form_history
            columns = [c['name'].lower() for c in inspector.get_columns('form_history')]
            if 'is_deleted' not in columns:
                if 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']:
                    db.session.execute(text('ALTER TABLE form_history ADD COLUMN is_deleted BOOLEAN DEFAULT 0'))
                else:
                    db.session.execute(text('ALTER TABLE form_history ADD COLUMN is_deleted BOOLEAN DEFAULT FALSE'))
                db.session.commit()
            if 'don_hang' not in columns:
                if 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']:
                    db.session.execute(text("ALTER TABLE form_history ADD COLUMN don_hang VARCHAR(255) DEFAULT ''"))
                else:
                    db.session.execute(text("ALTER TABLE form_history ADD COLUMN don_hang VARCHAR(255) DEFAULT ''"))
                db.session.commit()
            if 'nguoi_phu_trach' not in columns:
                if 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']:
                    db.session.execute(text("ALTER TABLE form_history ADD COLUMN nguoi_phu_trach VARCHAR(100) DEFAULT ''"))
                else:
                    db.session.execute(text("ALTER TABLE form_history ADD COLUMN nguoi_phu_trach VARCHAR(100) DEFAULT ''"))
                db.session.commit()
                print("Added nguoi_phu_trach column. Running data migration for NPT...")
                try:
                    import json
                    page = 0
                    per_page = 50
                    while True:
                        records = FormHistory.query.order_by(FormHistory.id).limit(per_page).offset(page * per_page).all()
                        if not records:
                            break
                        for r in records:
                            try:
                                jd = json.loads(r.data_json) if r.data_json else {}
                                f48 = (jd.get('f48') or '').strip()
                                if f48:
                                    final_npt = normalize_npt(f48)
                                    if final_npt != f48:
                                        jd['f48'] = final_npt
                                    r.nguoi_phu_trach = final_npt
                                    r.data_json = json.dumps(jd, ensure_ascii=False)
                            except Exception as parse_ex:
                                print(f"Error parsing json for record {r.id}: {parse_ex}")
                        db.session.commit()
                        page += 1
                    print("Finished NPT data migration.")
                except Exception as mig_ex:
                    print(f"Error during NPT data migration: {mig_ex}")
                    db.session.rollback()
        except Exception as ex:
            print(f"⚠️ Column migration skipped or failed: {ex}")
            db.session.rollback()

        # Ensure Employee table has initial data from FormHistory if empty
        try:
            if Employee.query.count() == 0:
                print("Populating initial Employee list from FormHistory...")
                unique_npts = db.session.query(FormHistory.nguoi_phu_trach).filter(FormHistory.nguoi_phu_trach != '').distinct().all()
                emps = set()
                for row in unique_npts:
                    val = row[0]
                    if not val: continue
                    emp_name = val.split(' - ')[0] if ' - ' in val else val
                    emps.add(emp_name.strip())
                for emp_name in sorted(emps):
                    if emp_name:
                        db.session.add(Employee(name=emp_name))
                db.session.commit()
                print(f"Added {len(emps)} employees.")
        except Exception as emp_ex:
            print(f"⚠️ Employee list init failed: {emp_ex}")
            db.session.rollback()
    except Exception as e:
        print(f"❌ Database initialization error: {e}")



BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- TRANSLATION MAPS ---
FIXED_TRANS = {
    'độc thân': '未婚', 'doc than': '未婚', 'đã kết hôn': '已婚', 'da ket hon': '已婚', 'có gia đình': '已婚',
    'ly hôn': '離婚', 'ly hon': '離婚', 'góa': '喪偶', 'goa': '喪偶',
    'tiểu học': '國小', 'tieu hoc': '國小', 'thcs': '國中', 'trung học cơ sở': '國中',
    'thpt': '高中', 'trung học phổ thông': '高中', 'trung cấp': '高職', 'trung cap': '高職',
    'cao đẳng': '專科', 'cao dang': '專科', 'đại học': '大學', 'dai hoc': '大學',
    'thạc sĩ': '碩士', 'thac si': '碩士', 'tiến sĩ': '博士', 'tien si': '博士',
    'việt nam': '越南', 'viet nam': '越南', 'đài loan': '台灣', 'nhật bản': '日本',
    'hàn quốc': '韓國', 'malaysia': '馬來西亞', 'macau': '澳門', 'thái lan': '泰國',
    'châu âu': '歐洲', 'nga': '俄羅斯',
    'nghệ an': '藝安', 'nghe an': '藝安',
    'may': '縫紉', 'thợ may': '縫紉', 'may mặc': '縫紉', 'hàn': '焊接', 'thợ hàn': '焊接',
    'điện': '電工', 'thợ điện': '電工', 'sơn': '噴漆', 'thợ sơn': '噴漆',
    'tiện': '車床', 'thợ tiện': '車床', 'phay': '銑床', 'thợ phay': '銑床',
    'bào': '刨床', 'thợ bào': '刨床', 'đúc': '鑄造', 'thợ đúc': '鑄造',
    'dệt': '紡織', 'thợ dệt': '紡織', 'mộc': '木工', 'thợ mộc': '木工',
    'in ấn': '印刷', 'in': '印刷', 'cơ khí': '機械', 'gia công': '加工',
    'xây dựng': '營造', 'xây': '營造', 'lắp ráp': '組裝', 'đóng gói': '包裝',
    'kiểm tra': '檢查', 'qc': '品檢', 'kho': '倉庫', 'thủ kho': '倉管',
    'nấu ăn': '烹飪', 'đầu bếp': '廚師', 'giúp việc': '幫傭', 'điều dưỡng': '護理',
    'chăm sóc người già': '照顧老人', 'nông nghiệp': '農業', 'nông': '農業',
    'chăn nuôi': '畜牧', 'thủy sản': '水產', 'nhựa': '塑膠',
    'lái xe': '駕駛', 'tài xế': '司機', 'xe nâng': '堆高機', 'lái xe nâng': '堆高機',
    'cnc': 'CNC', 'tig mig': 'Tig/Mig',
}


YELLOW_ALERTS_MAP = {
    "f05": "骨折", "f06": "手汗", "f11": "脊椎受傷", "f13": "伏地挺身 10~30 下",
    "f14": "搬重 20~40 kg", "f18": "肝炎", "f19": "斷指", "f20": "哮喘",
    "f21": "伏地挺身 50 下以上", "f22": "搬重 50kg 以上",
}

SKILL_MAPPING = {
    'f23': 'Hàn điện / 電焊', 'f24': 'Hàn Argon / 氬焊', 'f25': 'Hàn CO2 / 氣焊', 'f26': 'Tig/Mig',
    'f31': 'Tiện / 車床', 'f32': 'Phay / 銑床', 'f33': 'Bào / 刨床', 'f34': 'CNC',
    'f35': 'Đột dập / 沖床', 'f41': 'Sửa chữa máy / 機械維修', 'f49': 'Lắp ráp cơ khí / 機械組裝', 'f27': 'Đúc / 鑄造',
    'f50': 'Công nhân điện tử / 電子工', 'f39': 'Nhựa / 塑膠', 'f36': 'In ấn / 印刷', 'f28': 'Dệt / 紡織',
    'f29': 'May / 縫紉', 'f40': 'Xây dựng / 營造', 'f37': 'Thợ mộc / 木工', 'f38': 'Lái xe tải/khách / 卡車/客司機',
    'f30': 'Lái xe nâng / 堆高機', 'f44': 'Xe cẩu / 吊車', 'f45': 'Cẩu trục / 天車', 'f46': 'Máy xúc / 挖土機',
    'f42': 'Điều dưỡng / 護理工', 'f43': 'Giúp việc / 幫傭'
}

ZH_TO_VI = {
    # Học vấn
    "國中": "Cấp 2",
    "高中": "Cấp 3",
    "中專": "Trung cấp",
    "大專": "Cao đẳng",
    "大學": "Đại học",
    "國小": "Cấp 1",
    
    # Nơi ở / 63 Tỉnh Thành
    "萊州": "Lai Châu",
    "奠邊": "Điện Biên",
    "山羅": "Sơn La",
    "和平": "Hòa Bình",
    "安沛": "Yên Bái",
    "老街": "Lào Cai",
    "河江": "Hà Giang",
    "高平": "Cao Bằng",
    "北𣴓": "Bắc Kạn",
    "諒山": "Lạng Sơn",
    "宣光": "Tuyên Quang",
    "太原": "Thái Nguyên",
    "富壽": "Phú Thọ",
    "北江": "Bắc Giang",
    "廣寧": "Quảng Ninh",
    "北寧": "Bắc Ninh",
    "河南": "Hà Nam",
    "河內": "Hà Nội",
    "海陽": "Hải Dương",
    "海防": "Hải Phòng",
    "興安": "Hưng Yên",
    "南定": "Nam Định",
    "寧平": "Ninh Bình",
    "太平": "Thái Bình",
    "永福": "Vĩnh Phúc",
    "清化": "Thanh Hóa",
    "乂安": "Nghệ An",
    "藝安": "Nghệ An",
    "河靜": "Hà Tĩnh",
    "廣平": "Quảng Bình",
    "廣治": "Quảng Trị",
    "承天順化": "Thừa Thiên Huế",
    "峴港": "Đà Nẵng",
    "廣南": "Quảng Nam",
    "廣義": "Quảng Ngãi",
    "平定": "Bình Định",
    "富安": "Phú Yên",
    "慶和": "Khánh Hòa",
    "寧順": "Ninh Thuận",
    "平順": "Bình Thuận",
    "崑嵩": "Kon Tum",
    "嘉萊": "Gia Lai",
    "多樂": "Đắk Lắk",
    "得農": "Đắk Nông",
    "林同": "Lâm Đồng",
    "平福": "Bình Phước",
    "平陽": "Bình Dương",
    "同奈": "Đồng Nai",
    "西寧": "Tây Ninh",
    "巴地頭頓": "Bà Rịa - Vũng Tàu",
    "胡志明市": "TP Hồ Chí Minh",
    "隆安": "Long An",
    "同塔": "Đồng Tháp",
    "前江": "Tiền Giang",
    "安江": "An Giang",
    "檳椥": "Bến Tre",
    "永隆": "Vĩnh Long",
    "茶榮": "Trà Vinh",
    "後江": "Hậu Giang",
    "堅江": "Kiên Giang",
    "朔莊": "Sóc Trăng",
    "薄遼": "Bạc Liêu",
    "金甌": "Cà Mau",
    "芹苴": "Cần Thơ",
}


# --- HELPERS ---
def is_chinese(text: str) -> bool:
    if not text: return False
    cjk_count = sum(1 for c in text if '\u4e00' <= c <= '\u9fff' or '\u3400' <= c <= '\u4dbf' or '\uf900' <= c <= '\ufaff')
    non_space = len(text.replace(' ', '').replace(',', '').replace('，', '').replace('、', '').replace('。', ''))
    return non_space > 0 and (cjk_count / non_space) >= 0.5

def translate_fixed(text: str) -> str:
    if not text: return text
    return FIXED_TRANS.get(text.strip().lower(), text)

def translate_name(text: str) -> str:
    """Dành riêng cho dịch Họ Tên: Ưu tiên từ điển tên để tránh nhầm với tiếng Anh"""
    if not text or not text.strip() or is_chinese(text): return text

    # Chuẩn hóa Unicode NFC đầu vào để khớp chính xác từ điển nội bộ
    text_normalized = normalize('NFC', text).strip()

    # 1. Thử dịch từ từ điển tên riêng
    # Hàm trả về None nếu có phần nào không tìm thấy → fallback Google Translate
    dict_result = get_vietnamese_name_in_chinese(text_normalized)
    if dict_result is not None:
        return dict_result

    # 2. Nếu không có trong từ điển, dùng Gemini API hoặc Google Translate
    try:
        if gemini_api_key:
            model = genai.GenerativeModel('gemini-3.5-flash')
            prompt = f"Dịch tên tiếng Việt sau sang tiếng Trung Phồn Thể một cách tự nhiên nhất (âm Hán Việt nếu có thể), chỉ trả về đúng tên đã dịch, tuyệt đối không giải thích thêm: {text_normalized}"
            response = model.generate_content(prompt)
            result = response.text.strip()
            return result if result else text_normalized
        else:
            result = GoogleTranslator(source='vi', target='zh-TW').translate(text_normalized)
            return result if result else text_normalized
    except Exception as e:
        print(f"Name translation error: {e}")
        try:
            # Fallback to Google Translate if Gemini fails
            return GoogleTranslator(source='vi', target='zh-TW').translate(text_normalized) or text_normalized
        except:
            return text_normalized

def translate_free(text: str) -> str:
    """Dành cho dịch nội dung tự do (công việc, địa chỉ): Ưu tiên thuật ngữ nghề nghiệp"""
    if not text or not text.strip() or is_chinese(text): return text
    
    # Chuẩn hóa Unicode NFC để tránh lỗi so khớp do khác biệt Unicode encoding (NFC/NFD)
    text_normalized = normalize('NFC', text)
    text_lower = text_normalized.strip().lower()
    
    # 1. Kiểm tra khớp chính xác trong FIXED_TRANS
    fixed = FIXED_TRANS.get(text_lower)
    if fixed: return fixed
    
    # 2. Xử lý các từ khóa quan trọng TRONG câu (ví dụ: "may" -> "縫紉")
    # Để tránh Google dịch nhầm "may" thành "có lẽ/có thể" (possibly)
    processed_text = text_normalized
    # Các từ khóa cần bảo vệ (case insensitive)
    protected_terms = {
        'may': '縫紉',
        'thợ may': '縫紉',
        'may mặc': '縫紉',
        'thợ sơn': '噴漆',
        'thợ hàn': '焊接',
        'thợ điện': '電工',
        'lái xe': '駕駛',
        'nghệ an': '藝安',
    }
    
    # Nếu text là một từ đơn nằm trong danh sách bảo vệ, trả về luôn
    if text_lower in protected_terms:
        return protected_terms[text_lower]

    # 3. Dùng Gemini API hoặc Google Translate cho đoạn văn
    try:
        if gemini_api_key:
            model = genai.GenerativeModel('gemini-3.5-flash')
            prompt = f"Bạn là chuyên gia dịch thuật CV xuất khẩu lao động Đài Loan. Hãy dịch đoạn kinh nghiệm làm việc sau sang tiếng Trung Phồn Thể. Yêu cầu: dịch sát nghĩa, chuẩn thuật ngữ nghề nghiệp (cơ khí, điện, xây dựng, nhà máy, dệt may...), giữ nguyên cách dòng và định dạng nếu có. Tuyệt đối KHÔNG kèm theo lời giải thích hay bình luận, chỉ trả về đúng kết quả dịch. Đoạn văn bản cần dịch: '{processed_text.strip()}'"
            response = model.generate_content(prompt)
            result = response.text.strip()
        else:
            result = GoogleTranslator(source='vi', target='zh-TW').translate(processed_text.strip())
        
        # Sửa lại nếu Google dịch nhầm "may" -> "可能"
        if '可能' in result and 'may' in text_lower:
            result = result.replace('可能', '縫紉')
            
        # Chuẩn hóa bỏ dấu để kiểm tra Nghệ An chính xác (bao phủ NFC, NFD, không dấu)
        import unicodedata
        clean_text = ''.join(c for c in unicodedata.normalize('NFKD', text_normalized).lower() if not unicodedata.combining(c))
        if 'nghe an' in clean_text:
            # Thay thế tất cả các dạng dịch sai thường gặp (Giản thể/Phồn thể/乂安) sang 藝安
            for bad_trans in ('义安', '義安', '乂安'):
                if bad_trans in result:
                    result = result.replace(bad_trans, '藝安')
            # Thay thế cả chuỗi tiếng Việt/tiếng Anh chưa dịch được còn sót lại
            import re
            result = re.sub(r'(?i)nghệ\s+an', '藝安', result)
            result = re.sub(r'(?i)nghe\s+an', '藝安', result)
                    
        return result if result else text_normalized
    except Exception as e:
        print(f"Free text translation error: {e}")
        try:
            # Fallback to Google Translate if Gemini fails
            return GoogleTranslator(source='vi', target='zh-TW').translate(processed_text.strip()) or text_normalized
        except:
            return text_normalized

def sanitize_filename_master(name):
    if not name: return "UnNamed"
    # Xử lý riêng Đ/đ vì không phân rã được bằng NFKD → bị encode ASCII mất hoàn toàn
    s = str(name).replace('Đ', 'D').replace('đ', 'd')
    s = normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'[^\w\s-]', '', s).strip()
    s = re.sub(r'[-\s]+', '_', s)
    return s

def fmt_date(d):
    if not d: return ""
    try: return datetime.strptime(str(d), "%Y-%m-%d").strftime("%d/%m/%Y")
    except: return str(d)

def calc_age(birth_str):
    if not birth_str: return ""
    try:
        b = datetime.strptime(str(birth_str), "%Y-%m-%d")
        today = date.today()
        return today.year - b.year - ((today.month, today.day) < (b.month, b.day))
    except: return ""

def chk(v):
    return v in (True, 'true', '1', 1, 'yes', 'on', 'checked')

def get_base64_image(path, max_size=None, quality=80):
    if not os.path.exists(path): return ""
    try:
        if max_size:
            img = Image.open(path)
            img.thumbnail((max_size, max_size), Image.LANCZOS)
            if img.mode in ("RGBA", "P"): img = img.convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=quality)
            return f"data:image/jpeg;base64,{base64.b64encode(buf.getvalue()).decode('utf-8')}"
        with open(path, "rb") as f:
            ext = path.split('.')[-1].lower()
            return f"data:image/{ext};base64,{base64.b64encode(f.read()).decode('utf-8')}"
    except: return ""

# --- CACHE ---
_LOGO_B64_CACHE = None
_BG_B64_CACHE   = None
_TEMPLATE_CACHE = None

def _init_cache():
    global _LOGO_B64_CACHE, _BG_B64_CACHE, _TEMPLATE_CACHE
    _LOGO_B64_CACHE = get_base64_image(os.path.join(BASE_DIR, 'static', 'logo.png'))
    _BG_B64_CACHE   = get_base64_image(os.path.join(BASE_DIR, 'static', 'fct_bg.png'), max_size=400, quality=75)
    try:
        with open(os.path.join(BASE_DIR, 'templates', 'fct_template_v6.18.html'), 'r', encoding='utf-8') as f:
            _TEMPLATE_CACHE = f.read()
    except: pass

_init_cache()


def _fetch_r2_image_as_base64(url: str) -> str:
    """Tải ảnh từ URL R2 về và chuyển thành Base64 để nhúng vào HTML offline.
    Đây là cầu nối giữa R2 (lưu file thật) và file HTML (cần Base64 để offline được).
    """
    if not url or not url.startswith('http'):
        return ""
    try:
        resp = requests.get(url, timeout=10)
        if resp.status_code != 200:
            print(f"⚠️ R2 fetch failed ({resp.status_code}): {url}")
            return ""
        # Xác định mime type từ Content-Type header hoặc extension URL
        content_type = resp.headers.get('Content-Type', 'image/jpeg').split(';')[0].strip()
        b64 = base64.b64encode(resp.content).decode('utf-8')
        return f"data:{content_type};base64,{b64}"
    except Exception as e:
        print(f"⚠️ R2 fetch error: {e}")
        return ""

# --- LOGIC PREPARE ---
def prepare_render_data(raw_data: dict) -> dict:
    data = {}
    fields = [
        'Maso', 'Hoten', 'TentiengTrung', 'Ngaysinh', 'Tuoi', 'Chieucao', 'Cannang', 
        'Lienhe', 'Noio', 'HotenBo', 'TB', 'HotenMe', 'TM', 'VoChong', 'VC', 
        'Socon', 'Anhchiem', 'Xepthu', 'f48', 'N1', 'N2', 'N3', 'ndcv1', 'ndcv2', 'ndcv3', 
        'loi_binh_1', 'Honnhan', 'Hocvan', 'QG1', 'QG2', 'QG3', 'video_link_1', 'video_link_2'
    ]
    for f in fields:
        val = str(raw_data.get(f, '')).strip()
        if f in ['Honnhan', 'Hocvan']: data[f] = FIXED_TRANS.get(val.lower(), val)
        elif f in ['QG1', 'QG2', 'QG3']: data[f] = translate_fixed(val)
        else: data[f] = val

    # Viết hoa tên tiếng Việt ở tầng Python để giữ đầy đủ dấu
    # (CSS text-transform:uppercase làm mất dấu do lỗi glyph font)
    if data.get('Hoten'):
        data['Hoten'] = data['Hoten'].upper()

    data['Ngaysinh'] = fmt_date(data['Ngaysinh'])
    if not data['Tuoi'] and raw_data.get('Ngaysinh'): data['Tuoi'] = calc_age(raw_data.get('Ngaysinh'))

    fields_to_translate = ['Noio', 'ndcv1', 'ndcv2', 'ndcv3', 'loi_binh_1', 'N1', 'N2', 'N3']
    non_empty = {f: data[f] for f in fields_to_translate if data.get(f, '').strip()}
    if non_empty:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=min(len(non_empty), 4)) as executor:
            future_map = {executor.submit(translate_free, v): k for k, v in non_empty.items()}
            for future in as_completed(future_map):
                field = future_map[future]
                try: data[field] = future.result()
                except: pass

    yellow_alerts = []
    for i in range(1, 23):
        key = f'f{i:02d}'
        if chk(raw_data.get(key)) and key in YELLOW_ALERTS_MAP: yellow_alerts.append(YELLOW_ALERTS_MAP[key])
    if yellow_alerts:
        alert_str = "、".join(yellow_alerts)
        data['loi_binh_1'] = (data['loi_binh_1'] + "、" + alert_str) if data.get('loi_binh_1') else alert_str

    skills_html = []
    for key, name in SKILL_MAPPING.items():
        if chk(raw_data.get(key)):
            skills_html.append(f'<span class="skill-tag">{name}</span>')
    data['KyNangList_HTML'] = "".join(skills_html)

    tay = []
    if chk(raw_data.get('f01')): tay.append("右手")
    if chk(raw_data.get('f07')): tay.append("左手")
    data['TayThuan'] = " / ".join(tay) if tay else "右手"
    
    vision_map = [('f02', "右眼受損"), ('f08', "左眼受損"), ('f16', "散光"), ('f17', "色盲"), ('f15', "近視")]
    vision = [label for k, label in vision_map if chk(raw_data.get(k))]
    data['ThiLuc'] = " / ".join(vision) if vision else "正常"
    data['f12'] = "Có / 有" if chk(raw_data.get('f12')) else "無"

    # Hút thuốc / Uống rượu
    hut = "抽菸" if chk(raw_data.get('f03')) else ("不抽菸" if chk(raw_data.get('f09')) else "")
    ruou = "喝酒" if chk(raw_data.get('f04')) else ("不喝酒" if chk(raw_data.get('f10')) else "")
    parts = [p for p in [hut, ruou] if p]
    data['HutRuou'] = " / ".join(parts) if parts else "無"

    for key in ('photo', 'qr_line'):
        path = raw_data.get(key, '')
        if isinstance(path, str) and path.startswith('data:image/'):
            # Đã là Base64 (upload từ Local hoặc cũ) → dùng luôn
            data[f'{key}_base64'] = path
        elif isinstance(path, str) and path.startswith('http'):
            # Là URL từ Cloudflare R2 → Kéo về và chuyển thành Base64 để nhúng offline
            print(f"🔄 Fetching {key} from R2 for offline HTML...")
            data[f'{key}_base64'] = _fetch_r2_image_as_base64(path)
        else:
            data[f'{key}_base64'] = ""

    # Xử lý ảnh tài liệu (giấy tờ) để render trang 2
    data['document_images'] = raw_data.get('document_images', [])

    return data

def _protect_html(html: str) -> str:
    html = re.sub(r'<!--(?!\[if).*?-->', '', html, flags=re.DOTALL)
    html = re.sub(r'>\s+<', '><', html)
    html = re.sub(r'\s{2,}', ' ', html).strip()
    anti_devtools = (
        '<script>'
        '(function(){'
        'document.addEventListener("contextmenu",function(e){e.preventDefault();},false);'
        'document.addEventListener("keydown",function(e){'
        'if(e.keyCode===123||(e.ctrlKey&&e.shiftKey&&(e.keyCode===73||e.keyCode===74||e.keyCode===67))||(e.ctrlKey&&e.keyCode===85)){'
        'e.preventDefault();e.stopPropagation();return false;}'
        '});'
        # Cải thiện devtools detection: chỉ kích hoạt khi THỰC SỰ có DevTools
        # Kiểm tra cả width/height diff VÀ tỷ lệ zoom để tránh false positive
        'var _t=function(){'
        'var wDiff=window.outerWidth-window.innerWidth;'
        'var hDiff=window.outerHeight-window.innerHeight;'
        'var zoom=window.devicePixelRatio||1;'
        # Chỉ trigger khi diff lớn VÀ zoom gần 1 (không phải user zoom)
        'if((wDiff>400||hDiff>400)&&zoom>=0.8&&zoom<=1.2){'
        'document.body.innerHTML="";'
        '}'
        '};'
        'setInterval(_t,1000);'
        '})();'
        '</script>'
    )
    return html.replace('</body>', anti_devtools + '</body>')

def generate_html_resume(form_data: dict, template_name='fct_template_v6.18.html') -> str:
    processed_data = prepare_render_data(form_data)
    processed_data['logo_base64'] = _LOGO_B64_CACHE or get_base64_image(os.path.join(BASE_DIR, 'static', 'logo.png'))
    processed_data['bg_base64']   = _BG_B64_CACHE   or get_base64_image(os.path.join(BASE_DIR, 'static', 'fct_bg.png'), max_size=400, quality=75)
    
    # Nhúng dữ liệu gốc (không chứa ảnh base64 nặng) để có thể nạp lại sau này
    raw_for_embed = {k: v for k, v in form_data.items() if k not in ('photo', 'qr_line', 'document_images')}
    # Đánh dấu nếu có ảnh
    if form_data.get('photo'): raw_for_embed['__has_photo'] = True
    if form_data.get('qr_line'): raw_for_embed['__has_qr'] = True
    raw_json_str = json.dumps(raw_for_embed, ensure_ascii=False)
    
    # Dùng placeholder để _protect_html không phá hủy JSON
    processed_data['raw_data_json'] = '___FCT_RAW_PLACEHOLDER___'
    
    if _TEMPLATE_CACHE: template = Template(_TEMPLATE_CACHE)
    else:
        with open(os.path.join(BASE_DIR, 'templates', template_name), 'r', encoding='utf-8') as f:
            template = Template(f.read())
    html = _protect_html(template.render(processed_data))
    # Thay placeholder bằng JSON thật SAU KHI minify xong
    return html.replace('___FCT_RAW_PLACEHOLDER___', raw_json_str)

def _resize_image_for_db(data_uri: str, max_px: int = 1200, quality: int = 85) -> str:
    if not data_uri or not data_uri.startswith('data:image/'): return data_uri
    try:
        header, encoded = data_uri.split(',', 1)
        img_bytes = base64.b64decode(encoded)
        img = Image.open(io.BytesIO(img_bytes))
        if max(img.width, img.height) > max_px: img.thumbnail((max_px, max_px), Image.LANCZOS)
        buf = io.BytesIO()
        if img.mode in ('RGBA', 'P'): img = img.convert('RGB')
        img.save(buf, format='JPEG', quality=quality, optimize=True)
        return f"data:image/jpeg;base64,{base64.b64encode(buf.getvalue()).decode('utf-8')}"
    except: return data_uri

def _prepare_data_for_db(data: dict) -> dict:
    clean = dict(data)
    for key in ('photo', 'qr_line'):
        if clean.get(key): clean[key] = _resize_image_for_db(clean[key])
    # Resize ảnh tài liệu nếu có - Giảm sâu hơn để tránh lỗi 413 (900px, quality 60)
    if clean.get('document_images') and isinstance(clean['document_images'], list):
        clean['document_images'] = [_resize_image_for_db(img, max_px=900, quality=60) for img in clean['document_images'] if img]
    return clean

# --- API ROUTES ---
@app.route('/')
def user_form(): return render_template('user_form.html')

@app.route('/fct-1503')
@auth_required
def index():
    resp = make_response(render_template('index.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/api/health')
def health(): return jsonify({'ok': True, 'msg': 'DAS V3.0 running'})

def _process_form_data(request):
    if request.content_type and 'multipart/form-data' in request.content_type:
        data = json.loads(request.form.get('data', '{}'))
        for key in ('photo', 'qr_line'):
            file = request.files.get(key)
            if file:
                ext = os.path.splitext(file.filename)[1][1:].lower() or 'png'
                data[key] = f"data:image/{'jpeg' if ext=='jpg' else ext};base64,{base64.b64encode(file.read()).decode('utf-8')}"
    else: data = request.get_json() or {}
    return data

@app.route('/api/submit-only', methods=['POST'])
def api_submit_only():
    try:
        data = _process_form_data(request)
        record_id = data.get('_record_id')
        ma_so = str(data.get('Maso', '')).strip()
        ho_ten = str(data.get('Hoten', '')).strip()

        if ma_so and ma_so.upper() != 'CHO_DUYET':
            existing = FormHistory.query.filter_by(ma_so=ma_so, is_deleted=False).first()
            if existing and (not record_id or int(record_id) != existing.id):
                return jsonify({'success': False, 'error': f'Mã số "{ma_so}" đã tồn tại.'}), 400

        if record_id:
            record = FormHistory.query.get(int(record_id))
            if not record: return jsonify({'success': False, 'error': 'Not found'}), 404
            old_data = json.loads(record.data_json) if record.data_json else {}
            for key in ('photo', 'qr_line', 'document_images'):
                if key not in data and old_data.get(key): data[key] = old_data[key]
            record.ma_so = ma_so or 'CHO_DUYET'
            record.ho_ten = ho_ten
            
            # Xử lý NPT
            if 'f48_emp' in data or 'f48_partner' in data:
                emp = data.get('f48_emp', '').strip()
                partner = data.get('f48_partner', '').strip()
                if emp and partner: f48_raw = f"{emp} - {partner}"
                elif emp: f48_raw = emp
                elif partner: f48_raw = partner
                else: f48_raw = ""
                data['f48'] = f48_raw
            else:
                f48_raw = data.get('f48', '').strip()
            
            final_npt = normalize_npt(f48_raw)
            data['f48'] = final_npt
            record.nguoi_phu_trach = final_npt

            if 'Donhang' in data:
                record.don_hang = str(data.get('Donhang', '')).strip()
            else:
                data['Donhang'] = record.don_hang or ''
            record.data_json = json.dumps(_prepare_data_for_db(data), ensure_ascii=False)
            db.session.commit()
            msg = 'Đã cập nhật.'
        else:
            don_hang = str(data.get('Donhang', '')).strip()
            
            # Xử lý NPT
            if 'f48_emp' in data or 'f48_partner' in data:
                emp = data.get('f48_emp', '').strip()
                partner = data.get('f48_partner', '').strip()
                if emp and partner: f48_raw = f"{emp} - {partner}"
                elif emp: f48_raw = emp
                elif partner: f48_raw = partner
                else: f48_raw = ""
                data['f48'] = f48_raw
            else:
                f48_raw = data.get('f48', '').strip()
                
            final_npt = normalize_npt(f48_raw)
            data['f48'] = final_npt
            
            record = FormHistory(
                ma_so=ma_so or 'CHO_DUYET',
                ho_ten=ho_ten,
                don_hang=don_hang,
                nguoi_phu_trach=final_npt,
                data_json=json.dumps(_prepare_data_for_db(data), ensure_ascii=False)
            )
            db.session.add(record)
            db.session.commit()
            msg = 'Đã nộp form.'

        return jsonify({'success': True, 'id': record.id, 'ma_so': record.ma_so, 'msg': msg})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/translate', methods=['POST'])
def api_translate():
    try:
        data = request.get_json() or {}
        # Sử dụng translate_name cho API này vì nó thường được gọi cho ô Họ Tên
        return jsonify({'success': True, 'translated': translate_name(data.get('text', ''))})
    except: return jsonify({'success': False}), 500

@app.route('/api/download-cv/<maso>', methods=['GET'])
@auth_required
def download_history(maso):
    try:
        record = FormHistory.query.filter_by(ma_so=maso).order_by(FormHistory.ngay_tao.desc()).first()
        if not record: return jsonify({"error": "Not found"}), 404
        html_content = generate_html_resume(json.loads(record.data_json))
        filename = f"{maso}_{sanitize_filename_master(record.ho_ten)}.html"
        return send_file(io.BytesIO(html_content.encode('utf-8')), mimetype='text/html', as_attachment=True, download_name=filename)
    except: return jsonify({"error": "Error"}), 400

@app.route('/resume-<int:record_id>.html')
@auth_required
def api_preview(record_id):
    try:
        record = FormHistory.query.get(record_id)
        if not record: return "Not found", 404
        html_content = generate_html_resume(json.loads(record.data_json))
        return Response(html_content, mimetype="text/html", headers={"Content-Type": "text/html; charset=utf-8"})
    except Exception as e: return str(e), 500

@app.route('/cv/<path:slug>', methods=['GET'])
def secure_web_view(slug):
    try:
        # Hỗ trợ 2 định dạng:
        # 1. /cv/92/MD13243_Nguyen_Van_Chuc (id/slug) -> Ưu tiên ID (chính xác nhất)
        # 2. /cv/MD13243_Nguyen_Van_Chuc (slug) -> Tìm theo Maso (tiện lợi)
        
        parts = slug.split('/')
        record = None
        
        if len(parts) >= 2:
            # Định dạng id/slug
            try:
                rid = int(parts[0])
                record = FormHistory.query.get(rid)
            except: pass
        
        if not record:
            # Định dạng slug (tìm theo Maso)
            # Tách lấy phần trước dấu _ đầu tiên hoặc dùng cả slug nếu không có _
            maso_part = slug.split('_')[0] if '_' in slug else slug
            # Trường hợp đặc biệt: Nếu slug là số nguyên, thử tìm theo ID
            if slug.isdigit():
                record = FormHistory.query.get(int(slug))
            
            if not record:
                # Tìm bản ghi mới nhất có mã số này
                record = FormHistory.query.filter_by(ma_so=maso_part).order_by(FormHistory.ngay_tao.desc()).first()

        if not record: return "Không tìm thấy hồ sơ / Not found", 404
        
        # Kiểm tra maso trong slug (nếu có id/maso) để đảm bảo tính bảo mật/nhất quán
        # (Nếu dùng Maso từ slug thì record đã khớp rồi)
        
        html_content = generate_html_resume(json.loads(record.data_json))
        # Tạo tên file đẹp cho trình duyệt
        clean_name = sanitize_filename_master(record.ho_ten)
        filename = f"{record.ma_so}_{clean_name}.html"
        encoded_filename = quote(filename)
        
        response = Response(html_content, mimetype="text/html")
        response.headers["Content-Type"] = "text/html; charset=utf-8"
        response.headers["Content-Disposition"] = f'inline; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        return response
    except Exception as e:
        traceback.print_exc()
        return str(e), 500


@app.route('/api/employees', methods=['GET', 'POST', 'DELETE'])
@auth_required
def handle_employees():
    if request.method == 'GET':
        emps = Employee.query.order_by(Employee.name).all()
        return jsonify({'success': True, 'data': [e.name for e in emps]})
    elif request.method == 'POST':
        data = request.json or {}
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'success': False, 'message': 'Tên không hợp lệ'})
        existing = Employee.query.filter_by(name=name).first()
        if not existing:
            db.session.add(Employee(name=name))
            db.session.commit()
        return jsonify({'success': True, 'message': 'Đã thêm nhân viên'})
    elif request.method == 'DELETE':
        data = request.json or {}
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'success': False, 'message': 'Tên không hợp lệ'})
        emp = Employee.query.filter_by(name=name).first()
        if emp:
            db.session.delete(emp)
            db.session.commit()
            return jsonify({'success': True, 'message': 'Đã xóa nhân viên'})
        return jsonify({'success': False, 'message': 'Không tìm thấy nhân viên'})

@app.route('/api/history', methods=['GET'])
@auth_required
def api_history():
    try:
        from sqlalchemy.orm import load_only
        # load_only: chỉ SELECT các cột cần thiết, bỏ qua data_json (chứa ảnh base64 nặng hàng MB)
        records = (
            FormHistory.query
            .options(load_only(
                FormHistory.id,
                FormHistory.ma_so,
                FormHistory.ho_ten,
                FormHistory.ngay_tao,
                FormHistory.is_selected,
                FormHistory.don_hang,
                FormHistory.nguoi_phu_trach
            ))
            .filter_by(is_deleted=False)
            .order_by(FormHistory.ngay_tao.desc())
            .all()
        )
        vietnam_tz = timezone(timedelta(hours=7))
        data = [{
            'id': r.id, 'ma_so': r.ma_so, 'ho_ten': r.ho_ten,
            'don_hang': getattr(r, 'don_hang', '') or '',
            'is_selected': getattr(r, 'is_selected', False),
            'nguoi_phu_trach': getattr(r, 'nguoi_phu_trach', '') or '',
            'ngay_tao': r.ngay_tao.replace(tzinfo=timezone.utc).astimezone(vietnam_tz).strftime("%d/%m/%Y %H:%M:%S") if r.ngay_tao else ''
        } for r in records]
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/history/<int:record_id>/data', methods=['GET'])
@auth_required
def api_history_data(record_id):
    try:
        record = FormHistory.query.get(record_id)
        if not record: return jsonify({'success': False, 'error': 'Not found'}), 404
        return jsonify({
            'success': True, 
            'data_json': json.loads(record.data_json) if record.data_json else {}
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/history/bulk-delete', methods=['POST'])
@auth_required
def api_bulk_delete_history():
    try:
        data = request.get_json() or {}
        ids = data.get('ids', [])
        if not ids: return jsonify({'success': False, 'error': 'No IDs'}), 400
        records = FormHistory.query.filter(FormHistory.id.in_([int(i) for i in ids])).all()
        for r in records: 
            r.is_deleted = True
            try:
                if r.data_json:
                    jd = json.loads(r.data_json)
                    jd.pop('photo', None)
                    jd.pop('qr_line', None)
                    jd.pop('document_images', None)
                    r.data_json = json.dumps(jd)
            except: pass
        db.session.commit()
        return jsonify({'success': True, 'deleted': len(records)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/history/<int:record_id>', methods=['DELETE'])
@auth_required
def api_delete_history(record_id):
    try:
        r = FormHistory.query.get(record_id)
        if r:
            r.is_deleted = True
            try:
                if r.data_json:
                    jd = json.loads(r.data_json)
                    jd.pop('photo', None)
                    jd.pop('qr_line', None)
                    jd.pop('document_images', None)
                    r.data_json = json.dumps(jd)
            except: pass
            db.session.commit()
        return jsonify({'success': True})
    except: return jsonify({'success': False}), 500

@app.route('/api/history/hard-delete-year', methods=['POST'])
@auth_required
def api_hard_delete_year():
    try:
        data = request.get_json() or {}
        year_val = data.get('year')
        if not year_val or str(year_val) == 'ALL': return jsonify({'success': False, 'error': 'Invalid year'}), 400
        from sqlalchemy import extract
        records = FormHistory.query.filter(extract('year', FormHistory.ngay_tao) == int(year_val)).all()
        for r in records: db.session.delete(r)
        db.session.commit()
        return jsonify({'success': True, 'deleted': len(records)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/history/bulk-download', methods=['POST'])
@auth_required
def api_bulk_download():
    try:
        data = request.get_json() or {}
        ids = data.get('ids', [])
        if not ids: return jsonify({'success': False, 'error': 'No IDs'}), 400
        records = FormHistory.query.filter(FormHistory.id.in_([int(i) for i in ids])).all()
        if not records: return jsonify({'success': False, 'error': 'No records found'}), 404
        
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for r in records:
                try:
                    form_data = json.loads(r.data_json)
                    html_content = generate_html_resume(form_data)
                    filename = f"{r.ma_so}_{sanitize_filename_master(r.ho_ten)}.html"
                    zf.writestr(filename, html_content)
                except: pass
        zip_buffer.seek(0)
        return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name='FCT_HoSo_Export.zip')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/history/bulk-assign-job', methods=['POST'])
@auth_required
def api_bulk_assign_job():
    try:
        data = request.get_json() or {}
        ids = data.get('ids', [])
        don_hang = str(data.get('don_hang', '')).strip()
        if not ids: return jsonify({'success': False, 'error': 'No IDs provided'}), 400
        
        records = FormHistory.query.filter(FormHistory.id.in_([int(i) for i in ids])).all()
        for r in records:
            current_jobs_str = (r.don_hang or "").strip()
            new_donhang = don_hang
            
            if current_jobs_str:
                current_jobs = [j.strip() for j in re.split(r'[,;]+', current_jobs_str) if j.strip()]
                target_upper = unicodedata.normalize('NFC', don_hang.upper())
                already_has = any(unicodedata.normalize('NFC', j.upper()) == target_upper for j in current_jobs)
                
                if not already_has:
                    current_jobs.append(don_hang)
                    new_donhang = ", ".join(current_jobs)
                else:
                    new_donhang = current_jobs_str # giữ nguyên
                    
            r.don_hang = new_donhang
            try:
                if r.data_json:
                    jd = json.loads(r.data_json)
                    jd['Donhang'] = new_donhang
                    r.data_json = json.dumps(jd, ensure_ascii=False)
            except:
                pass
        db.session.commit()
        return jsonify({'success': True, 'count': len(records)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/history/assign-job-by-maso', methods=['POST'])
@auth_required
def api_assign_job_by_maso():
    try:
        data = request.get_json() or {}
        maso_list = data.get('maso_list', [])
        don_hang = str(data.get('don_hang', '')).strip()
        
        # Chuẩn hóa danh sách mã số và tên đơn hàng mục tiêu
        clean_masos = [unicodedata.normalize('NFC', str(m).strip().upper()) for m in maso_list if str(m).strip()]
        target_job = unicodedata.normalize('NFC', don_hang.strip().upper())
        
        # Nếu Tên đơn hàng trống, ta thực hiện xóa sạch đơn hàng cho các mã số trong list
        if not target_job:
            if not clean_masos:
                return jsonify({'success': False, 'error': 'No candidate codes provided'}), 400
            records = FormHistory.query.filter(func.upper(FormHistory.ma_so).in_(clean_masos)).all()
            for r in records:
                r.don_hang = ""
                try:
                    if r.data_json:
                        jd = json.loads(r.data_json)
                        jd['Donhang'] = ""
                        r.data_json = json.dumps(jd, ensure_ascii=False)
                except:
                    pass
            db.session.commit()
            return jsonify({'success': True, 'count': len(records)})
            
        # Tìm tất cả hồ sơ trong database để thực hiện gán mới và gỡ bỏ đồng bộ
        all_records = FormHistory.query.all()
        updated_count = 0
        
        for r in all_records:
            r_maso = (r.ma_so or '').strip().upper()
            if not r_maso:
                continue
                
            donhang_str = (r.don_hang or '').strip()
            # Tách các đơn hàng hiện tại của ứng viên
            current_jobs = [j.strip() for j in re.split(r'[,;]+', donhang_str) if j.strip()]
            current_jobs_upper = [unicodedata.normalize('NFC', j.upper()) for j in current_jobs]
            
            in_new_list = unicodedata.normalize('NFC', r_maso) in clean_masos
            has_job_currently = target_job in current_jobs_upper
            
            if in_new_list:
                # Nếu ứng viên có trong danh sách nhập vào nhưng chưa được gán đơn này, tiến hành gán thêm
                if not has_job_currently:
                    current_jobs.append(don_hang)
                    new_donhang = ", ".join(current_jobs)
                    r.don_hang = new_donhang
                    try:
                        if r.data_json:
                            jd = json.loads(r.data_json)
                            jd['Donhang'] = new_donhang
                            r.data_json = json.dumps(jd, ensure_ascii=False)
                    except:
                        pass
                    updated_count += 1
            else:
                # Nếu ứng viên KHÔNG có trong danh sách nhập vào nhưng hiện đang được gán đơn này, tiến hành gỡ bỏ
                if has_job_currently:
                    new_jobs = [j for j in current_jobs if unicodedata.normalize('NFC', j.strip().upper()) != target_job]
                    new_donhang = ", ".join(new_jobs)
                    r.don_hang = new_donhang
                    try:
                        if r.data_json:
                            jd = json.loads(r.data_json)
                            jd['Donhang'] = new_donhang
                            r.data_json = json.dumps(jd, ensure_ascii=False)
                    except:
                        pass
                    updated_count += 1
                    
        db.session.commit()
        return jsonify({'success': True, 'count': updated_count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/history/remove-job-from-maso', methods=['POST'])
@auth_required
def api_remove_job_from_maso():
    try:
        data = request.get_json() or {}
        ma_so = unicodedata.normalize('NFC', str(data.get('ma_so', '')).strip().upper())
        job_to_remove = unicodedata.normalize('NFC', str(data.get('don_hang', '')).strip().upper())
        if not ma_so or not job_to_remove:
            return jsonify({'success': False, 'error': 'Missing parameters'}), 400
            
        records = FormHistory.query.filter(func.upper(FormHistory.ma_so) == ma_so).all()
        updated_count = 0
        
        for r in records:
            donhang_str = (r.don_hang or '').strip()
            if not donhang_str:
                continue
            
            jobs = [j.strip() for j in re.split(r'[,;]+', donhang_str) if j.strip()]
            new_jobs = [j for j in jobs if unicodedata.normalize('NFC', j.strip().upper()) != job_to_remove]
            
            if len(new_jobs) != len(jobs):
                new_donhang = ", ".join(new_jobs)
                r.don_hang = new_donhang
                try:
                    if r.data_json:
                        jd = json.loads(r.data_json)
                        jd['Donhang'] = new_donhang
                        r.data_json = json.dumps(jd, ensure_ascii=False)
                except:
                    pass
                updated_count += 1
                
        db.session.commit()
        return jsonify({'success': True, 'count': updated_count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


EXCEL_SKILL_MAPPING = {
    'f23': 'Hàn điện', 'f24': 'Hàn argon', 'f25': 'Hàn CO2', 'f26': 'Tig Mig',
    'f31': 'Tiện', 'f32': 'Phay', 'f33': 'Bào', 'f34': 'CNC',
    'f35': 'Đột dập', 'f41': 'Sửa chữa máy', 'f49': 'Lắp ráp cơ khí', 'f27': 'Đúc',
    'f50': 'Công nhân điện tử', 'f39': 'Nhựa', 'f36': 'In ấn', 'f28': 'Dệt',
    'f29': 'May', 'f40': 'Xây dựng', 'f37': 'Thợ mộc', 'f38': 'Lái xe tải/khách',
    'f30': 'Xe nâng', 'f44': 'Xe cẩu', 'f45': 'Cẩu trục', 'f46': 'Máy xúc',
    'f42': 'Điều dưỡng', 'f43': 'Giúp việc'
}

@app.route('/api/history/export-excel', methods=['POST'])
@auth_required
def api_export_excel():
    try:
        data = request.get_json() or {}
        ids = data.get('ids', [])
        year_val = data.get('year')
        
        if ids:
            records = FormHistory.query.filter(FormHistory.id.in_([int(i) for i in ids])).all()
        elif year_val and str(year_val) != 'ALL':
            records = FormHistory.query.filter(func.extract('year', FormHistory.ngay_tao) == int(year_val)).all()
        else:
            records = FormHistory.query.all()
            
        if not records:
            return jsonify({'success': False, 'error': 'No records found'}), 404
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách ứng viên"
        
        # 1. Main Title
        ws.merge_cells("A1:K1")
        title_cell = ws.cell(row=1, column=1, value="BÁO CÁO TỔNG HỢP DANH SÁCH ỨNG VIÊN - FCT HUMAN RESOURCE")
        title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # 2. Subtitle
        ws.merge_cells("A2:K2")
        from datetime import datetime
        export_time_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        subtitle_cell = ws.cell(row=2, column=1, value=f"Thời gian xuất báo cáo: {export_time_str}   |   Tổng số ứng viên: {len(records)} hồ sơ")
        subtitle_cell.font = Font(name="Segoe UI", size=11, italic=True, color="1E3A8A")
        subtitle_cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 25

        # Spacer row 3
        ws.append([""] * 11)
        ws.row_dimensions[3].height = 10

        # Table headers at Row 4
        headers = ['Mã số', 'Họ tên', 'Ngày sinh', 'Trạng thái', 'Chiều cao (cm)', 'Cân nặng (kg)', 
                   'Trình độ văn hóa', 'Nơi ở', 'Tay nghề', 'Kinh nghiệm công việc', 'Người phụ trách']
        ws.append(headers)
        
        # Thống kê
        skills_count = {}
        edu_count = {}
        location_count = {}
        type_count = {'Nam (MD)': 0, 'Nữ (FD)': 0, 'Điều dưỡng (KD)': 0, 'Khác': 0}
        job_type_count = {'Nam công xưởng': 0, 'Nữ công xưởng': 0, 'Giúp việc & Khác': 0, 'Điều dưỡng': 0}
        
        for r in records:
            try:
                form_data = json.loads(r.data_json)
                ma_so = r.ma_so or ''
                ma_so_upper = ma_so.strip().upper()
                if ma_so_upper.startswith('MD'):
                    type_count['Nam (MD)'] += 1
                elif ma_so_upper.startswith('FD'):
                    type_count['Nữ (FD)'] += 1
                elif ma_so_upper.startswith('KD'):
                    type_count['Điều dưỡng (KD)'] += 1
                else:
                    type_count['Khác'] += 1
                    
                # Phân loại nhóm ngành nghề
                is_giup_viec = chk(form_data.get('f43')) # f43: Giúp việc
                if ma_so_upper.startswith('MD'):
                    job_type_count['Nam công xưởng'] += 1
                elif ma_so_upper.startswith('KD'):
                    job_type_count['Điều dưỡng'] += 1
                elif is_giup_viec:
                    job_type_count['Giúp việc & Khác'] += 1
                elif ma_so_upper.startswith('FD'):
                    job_type_count['Nữ công xưởng'] += 1
                else:
                    job_type_count['Giúp việc & Khác'] += 1
                    
                ho_ten = r.ho_ten or ''
                ngay_sinh = form_data.get('Ngaysinh', '')
                chieu_cao = form_data.get('Chieucao', '')
                can_nang = form_data.get('Cannang', '')
                hoc_van = form_data.get('Hocvan', '')
                noi_o = form_data.get('Noio', '')
                
                # Dịch học lực và nơi ở sang tiếng Việt cho báo cáo Excel
                hoc_van_vi = ZH_TO_VI.get(hoc_van.strip(), hoc_van) if hoc_van else ''
                noi_o_vi = ZH_TO_VI.get(noi_o.strip(), noi_o) if noi_o else ''
                
                nguoi_pt = form_data.get('f48', '')
                
                if hoc_van_vi:
                    edu_count[hoc_van_vi] = edu_count.get(hoc_van_vi, 0) + 1
                    
                if noi_o_vi:
                    noi_o_clean = noi_o_vi.strip()
                    if noi_o_clean:
                        location_count[noi_o_clean] = location_count.get(noi_o_clean, 0) + 1
                
                # Tay nghề
                skills = []
                for k, v in EXCEL_SKILL_MAPPING.items():
                    if form_data.get(k):
                        skills.append(v)
                        skills_count[v] = skills_count.get(v, 0) + 1
                tay_nghe = ", ".join(skills)
                
                # Kinh nghiệm
                kn = []
                for i in range(1, 4):
                    qg = form_data.get(f'QG{i}', '')
                    nam = form_data.get(f'N{i}', '')
                    cv = form_data.get(f'ndcv{i}', '')
                    if qg or cv:
                        parts = []
                        if qg: parts.append(qg)
                        if nam: parts.append(f"({nam})")
                        if cv: parts.append(f": {cv}")
                        kn.append(" ".join(parts))
                kinh_nghiem = "\n".join(kn)
                
                # Trạng thái
                trang_thai = '🎯 Trúng tuyển' if getattr(r, 'is_selected', False) else '📝 Gửi form'
                
                ws.append([ma_so, ho_ten, ngay_sinh, trang_thai, chieu_cao, can_nang, hoc_van_vi, noi_o_vi, tay_nghe, kinh_nghiem, nguoi_pt])
            except Exception as e:
                print("Error parsing record:", e)
                pass
                
        # Premium Styling
        thin_border = Border(left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"), top=Side(style='thin', color="CBD5E1"), bottom=Side(style='thin', color="CBD5E1"))
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # FCT Blue
        header_font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=11)
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Slate-50
        body_font = Font(name="Segoe UI", size=10)
        
        ws.row_dimensions[4].height = 30 # Header height
        ws.freeze_panes = 'A5'
        
        for col_num, cell in enumerate(ws[4], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        selected_row_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Xanh lá pastel (Emerald-100)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=11), 5):
            is_even = (row_idx % 2 == 0)
            ws.row_dimensions[row_idx].height = 40 # Tăng chiều cao mặc định cho dòng để dễ đọc hơn
            
            is_selected_row = (row[3].value == '🎯 Trúng tuyển')
            row_fill = selected_row_fill if is_selected_row else (zebra_fill if is_even else None)
            
            for cell in row:
                cell.font = body_font
                cell.border = thin_border
                # Column Kinh nghiệm (cột 10) căn trái, các cột khác căn giữa
                if cell.column == 10:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif cell.column == 9:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True) # Tay nghề căn trái
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                if row_fill:
                    cell.fill = row_fill
                    
        # Fixed professional column widths
        col_widths = {'A':15, 'B':26, 'C':15, 'D':20, 'E':16, 'F':16, 'G':20, 'H':20, 'I':32, 'J':55, 'K':22}
        for col_let, width in col_widths.items():
            ws.column_dimensions[col_let].width = width
            
        # Thêm AutoFilter
        ws.auto_filter.ref = f"A4:K{ws.max_row}"
            
        # Thêm Sheet Thống Kê
        ws_stat = wb.create_sheet(title="Thống Kê")
        
        # 1. Ẩn Gridlines - Tạo hiệu ứng phẳng như một trang Web Canvas cao cấp
        ws_stat.sheet_view.showGridLines = False
        
        # Font chữ mặc định hệ thống cao cấp Segoe UI
        font_family = "Segoe UI"
        
        # 2. Executive Title Banner (Cột D đến N)
        ws_stat.merge_cells("D1:N1")
        title_cell = ws_stat.cell(row=1, column=4, value="BÁO CÁO PHÂN TÍCH & THỐNG KÊ HỒ SƠ ỨNG VIÊN")
        title_cell.font = Font(name=font_family, size=14, bold=True, color="1E3A8A")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_stat.row_dimensions[1].height = 25
        
        ws_stat.merge_cells("D2:N2")
        subtitle_cell = ws_stat.cell(row=2, column=4, value="Hệ thống Quản lý Nhân sự FCT Human Resource  |  Báo cáo trực quan tự động")
        subtitle_cell.font = Font(name=font_family, size=9.5, italic=True, color="64748B") # Slate-500
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_stat.row_dimensions[2].height = 18
        
        # Dải ngăn cách thanh lịch (Row 3, tô nền màu Navy rất mỏng)
        ws_stat.merge_cells("D3:N3")
        separator_cell = ws_stat.cell(row=3, column=4)
        separator_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        ws_stat.row_dimensions[3].height = 2
        
        # 3. KPI Cards dạng Widget Web Cao cấp (Hàng 5 đến 7)
        kpi_card_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        border_left_thick = Border(left=Side(style='medium', color="1E3A8A"),
                                   right=Side(style='thin', color="E2E8F0"),
                                   top=Side(style='thin', color="E2E8F0"),
                                   bottom=Side(style='thin', color="E2E8F0"))
        
        border_right_only = Border(left=Side(style=None),
                                   right=Side(style='thin', color="E2E8F0"),
                                   top=Side(style='thin', color="E2E8F0"),
                                   bottom=Side(style='thin', color="E2E8F0"))
                                   
        def style_kpi_card(col_start, label, value):
            ws_stat.merge_cells(start_row=5, start_column=col_start, end_row=7, end_column=col_start+1)
            main_cell = ws_stat.cell(row=5, column=col_start, value=f"{label}\n\n{value}")
            main_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for r in range(5, 8):
                for c in range(col_start, col_start + 2):
                    cell = ws_stat.cell(row=r, column=c)
                    cell.fill = kpi_card_fill
                    if c == col_start:
                        cell.border = border_left_thick
                    else:
                        cell.border = border_right_only
            
            main_cell.font = Font(name=font_family, bold=True, size=10, color="1E3A8A")
            
        style_kpi_card(4, "TỔNG ỨNG VIÊN", f"{len(records)}")
        style_kpi_card(6, "NAM (MD)", f"{type_count['Nam (MD)']}")
        style_kpi_card(8, "NỮ (FD)", f"{type_count['Nữ (FD)']}")
        style_kpi_card(10, "ĐIỀU DƯỠNG (KD)", f"{type_count['Điều dưỡng (KD)']}")
        style_kpi_card(12, "MÃ KHÁC", f"{type_count['Khác']}")
        
        ws_stat.row_dimensions[5].height = 15
        ws_stat.row_dimensions[6].height = 15
        ws_stat.row_dimensions[7].height = 15
        
        # 4. Dynamic Data Tables in Columns A & B (Báo cáo kiểm toán)
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Slate-900 sang trọng
        header_font = Font(name=font_family, color="FFFFFF", bold=True, size=10)
        
        body_font = Font(name=font_family, size=10)
        body_bold_font = Font(name=font_family, bold=True, size=10)
        
        thin_border = Border(left=Side(style='thin', color="E2E8F0"), 
                             right=Side(style='thin', color="E2E8F0"), 
                             top=Side(style='thin', color="E2E8F0"), 
                             bottom=Side(style='thin', color="E2E8F0"))
                             
        double_bottom_border = Border(left=Side(style='thin', color="E2E8F0"),
                                      right=Side(style='thin', color="E2E8F0"),
                                      top=Side(style='thin', color="E2E8F0"),
                                      bottom=Side(style='double', color="000000")) # Đường gạch chân kép kiểm toán
                                      
        section_font = Font(name=font_family, bold=True, size=11, color="1E3A8A")
        
        header_rows = {}
        section_rows = {}
        total_rows = {}
        
        # Subsection I: Tay nghề
        row_skill = 9
        section_rows[row_skill - 1] = "I. THỐNG KÊ TAY NGHỀ & KỸ NĂNG"
        header_rows[row_skill] = ("Tay nghề", "Số lượng")
        
        for idx, (skill, count) in enumerate(skills_count.items(), 1):
            ws_stat.cell(row=row_skill + idx, column=1, value=skill)
            ws_stat.cell(row=row_skill + idx, column=2, value=count)
            
        row_skill_total = row_skill + len(skills_count) + 1
        total_rows[row_skill_total] = f"=SUM(B{row_skill+1}:B{row_skill_total-1})"
        
        # Subsection II: Nhóm ngành nghề
        row_job = row_skill_total + 3
        section_rows[row_job - 1] = "II. PHÂN BỔ NHÓM NGÀNH NGHỀ"
        header_rows[row_job] = ("Nhóm ngành nghề", "Số lượng")
        
        for idx, (job, count) in enumerate(job_type_count.items(), 1):
            ws_stat.cell(row=row_job + idx, column=1, value=job)
            ws_stat.cell(row=row_job + idx, column=2, value=count)
            
        row_job_total = row_job + len(job_type_count) + 1
        total_rows[row_job_total] = f"=SUM(B{row_job+1}:B{row_job_total-1})"
        
        # Subsection III: Trình độ văn hóa
        row_edu = row_job_total + 3
        section_rows[row_edu - 1] = "III. PHÂN BỔ TRÌNH ĐỘ VĂN HÓA"
        header_rows[row_edu] = ("Trình độ văn hóa", "Số lượng")
        
        for idx, (edu, count) in enumerate(edu_count.items(), 1):
            ws_stat.cell(row=row_edu + idx, column=1, value=edu)
            ws_stat.cell(row=row_edu + idx, column=2, value=count)
            
        row_edu_total = row_edu + len(edu_count) + 1
        total_rows[row_edu_total] = f"=SUM(B{row_edu+1}:B{row_edu_total-1})"
        
        # Subsection IV: Nơi ở / Quê quán
        row_loc = row_edu_total + 3
        section_rows[row_loc - 1] = "IV. PHÂN BỔ THEO NƠI Ở / QUÊ QUÁN"
        header_rows[row_loc] = ("Nơi ở / Quê quán", "Số lượng")
        
        sorted_locations = sorted(location_count.items(), key=lambda x: x[1], reverse=True)
        for idx, (loc, count) in enumerate(sorted_locations, 1):
            ws_stat.cell(row=row_loc + idx, column=1, value=loc)
            ws_stat.cell(row=row_loc + idx, column=2, value=count)
            
        row_loc_total = row_loc + len(sorted_locations) + 1
        total_rows[row_loc_total] = f"=SUM(B{row_loc+1}:B{row_loc_total-1})"
        
        # Ghi các ô Subsection, Header, Dòng Tổng Cộng và định dạng
        for r, title in section_rows.items():
            cell = ws_stat.cell(row=r, column=1, value=title)
            cell.font = section_font
            ws_stat.row_dimensions[r].height = 25
            
        for r, (h1, h2) in header_rows.items():
            cell1 = ws_stat.cell(row=r, column=1, value=h1)
            cell2 = ws_stat.cell(row=r, column=2, value=h2)
            for cell in [cell1, cell2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_stat.row_dimensions[r].height = 22
            
        for r, formula in total_rows.items():
            cell1 = ws_stat.cell(row=r, column=1, value="Tổng cộng")
            cell2 = ws_stat.cell(row=r, column=2, value=formula)
            for cell in [cell1, cell2]:
                cell.font = body_bold_font
                cell.border = double_bottom_border
            cell2.alignment = Alignment(horizontal="center", vertical="center")
            ws_stat.row_dimensions[r].height = 20
            
        all_headers = set(header_rows.keys())
        all_sections = set(section_rows.keys())
        all_totals = set(total_rows.keys())
        
        for r_idx in range(9, ws_stat.max_row + 1):
            if r_idx in all_headers or r_idx in all_sections or r_idx in all_totals:
                continue
            
            c1 = ws_stat.cell(row=r_idx, column=1)
            c2 = ws_stat.cell(row=r_idx, column=2)
            
            if c1.value is not None:
                ws_stat.row_dimensions[r_idx].height = 20
                is_even = (r_idx % 2 == 0)
                row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if is_even else None
                
                for cell in [c1, c2]:
                    cell.font = body_font
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill
                        
                c1.alignment = Alignment(horizontal="left", vertical="center")
                c2.alignment = Alignment(horizontal="center", vertical="center")

        # 5. Charts in Columns D to N (Segoe UI & Borderless Web View)
        
        # Chart 1: Tay nghề (Col BarChart)
        if skills_count:
            chart1 = BarChart()
            chart1.type = "col"
            chart1.style = 10
            chart1.title = "Thống kê Tay nghề (Kỹ năng)"
            chart1.y_axis.title = 'Số lượng ứng viên'
            chart1.x_axis.title = 'Tay nghề'
            chart1.width = 17
            chart1.height = 7.5
            
            data1 = Reference(ws_stat, min_col=2, min_row=row_skill, max_row=row_skill_total-1, max_col=2)
            cats1 = Reference(ws_stat, min_col=1, min_row=row_skill+1, max_row=row_skill_total-1)
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(cats1)
            chart1.legend = None
            chart1.shape = 4
            
            if chart1.series:
                series = chart1.series[0]
                series.dLbls = DataLabelList()
                series.dLbls.showVal = True
                
            ws_stat.add_chart(chart1, "D9")
            
        # Chart 2: Nhóm ngành nghề (PieChart)
        if job_type_count:
            chart_job = PieChart()
            chart_job.title = "Phân bổ Nhóm ngành nghề"
            chart_job.width = 17
            chart_job.height = 7.5
            chart_job.legend.position = "b"
            
            data_job = Reference(ws_stat, min_col=2, min_row=row_job, max_row=row_job_total-1)
            labels_job = Reference(ws_stat, min_col=1, min_row=row_job+1, max_row=row_job_total-1)
            chart_job.add_data(data_job, titles_from_data=True)
            chart_job.set_categories(labels_job)
            
            chart_job.dataLabels = DataLabelList()
            chart_job.dataLabels.showVal = True
            
            ws_stat.add_chart(chart_job, "J9")
            
        # Chart 3: Trình độ văn hóa (PieChart)
        if edu_count:
            pie = PieChart()
            pie.title = "Phân bổ Trình độ văn hóa"
            pie.width = 17
            pie.height = 7.5
            pie.legend.position = "b"
            
            data_edu = Reference(ws_stat, min_col=2, min_row=row_edu, max_row=row_edu_total-1)
            labels_edu = Reference(ws_stat, min_col=1, min_row=row_edu+1, max_row=row_edu_total-1)
            pie.add_data(data_edu, titles_from_data=True)
            pie.set_categories(labels_edu)
            
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showVal = True
            
            ws_stat.add_chart(pie, "D23")
            
        # Chart 4: Nơi ở / Quê quán (PieChart SIÊU TO KHỔNG LỒ)
        if location_count:
            chart_loc = PieChart()
            chart_loc.title = "Phân bổ theo Nơi ở / Quê quán"
            chart_loc.width = 24
            chart_loc.height = 11
            chart_loc.legend.position = "b"
            
            data_loc = Reference(ws_stat, min_col=2, min_row=row_loc, max_row=row_loc_total-1)
            cats_loc = Reference(ws_stat, min_col=1, min_row=row_loc+1, max_row=row_loc_total-1)
            chart_loc.add_data(data_loc, titles_from_data=True)
            chart_loc.set_categories(cats_loc)
            
            chart_loc.dataLabels = DataLabelList()
            chart_loc.dataLabels.showVal = True
            
            ws_stat.add_chart(chart_loc, f"D{row_loc}")

        # Cấu hình kích thước cột cho lưới
        ws_stat.column_dimensions['A'].width = 25
        ws_stat.column_dimensions['B'].width = 12
        ws_stat.column_dimensions['C'].width = 3
        
        for col_let in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws_stat.column_dimensions[col_let].width = 12
            
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_prefix = f"Nam{year_val}_" if year_val and str(year_val) != 'ALL' else ""
        return send_file(excel_buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                         as_attachment=True, download_name=f'FCT_UngVien_{filename_prefix}{timestamp}.xlsx')
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/history/bulk-print', methods=['POST'])
@auth_required
def api_bulk_print():
    try:
        data = request.get_json() or {}
        ids = data.get('ids', [])
        if not ids: return jsonify({'success': False, 'error': 'No IDs'}), 400
        records = FormHistory.query.filter(FormHistory.id.in_([int(i) for i in ids])).all()
        if not records: return jsonify({'success': False, 'error': 'No records found'}), 404
        
        # Sắp xếp đúng theo thứ tự ID được gửi lên từ Frontend
        records_dict = {r.id: r for r in records}
        sorted_records = [records_dict[int(i)] for i in ids if int(i) in records_dict]
        if not sorted_records: return jsonify({'success': False, 'error': 'No valid records'}), 404

        from flask import stream_with_context
        
        def generate_bulk_stream():
            # Bước 1: Render file hồ sơ gốc (để lấy phần <head> và các CSS, JS toàn cục)
            r0 = sorted_records[0]
            try:
                form_data = json.loads(r0.data_json)
                base_html = generate_html_resume(form_data)
            except Exception as e:
                yield f"<h1>Lỗi tạo trang đầu tiên: {str(e)}</h1>"
                return
                
            base_html = base_html.replace('page-break-after: avoid !important;', 'page-break-after: always !important;')
            
            # Tìm vị trí chèn các file tiếp theo (ngay trước thẻ <script> cuối cùng)
            insert_idx = base_html.find('<script id="fct-raw-data"')
            if insert_idx == -1:
                insert_idx = base_html.find('<script>(function()')
            if insert_idx == -1:
                insert_idx = base_html.find('<script>(function(){')
            if insert_idx == -1:
                # Fallback: chèn trước body
                insert_idx = base_html.rfind('</body>')
                
            if insert_idx != -1:
                # Gửi ngay phần đầu của trang (từ <!DOCTYPE> đến trước script)
                yield base_html[:insert_idx]
            else:
                yield base_html
                return
                
            # Bước 2: Streaming từng file hồ sơ tiếp theo mà không lưu toàn bộ vào RAM
            for r in sorted_records[1:]:
                try:
                    form_data = json.loads(r.data_json)
                    html = generate_html_resume(form_data)
                    
                    # Cắt lấy nguyên cụm thẻ <div class="a4-page ...">
                    start_idx = html.find('<div class="a4-page')
                    if start_idx != -1:
                        end_idx = html.find('<script id="fct-raw-data"', start_idx)
                        if end_idx == -1: end_idx = html.find('<script>(function()', start_idx)
                        if end_idx == -1: end_idx = html.find('<script>(function(){', start_idx)
                        if end_idx == -1: end_idx = html.rfind('</body>')
                            
                        if end_idx != -1:
                            # Stream phần thân của resume
                            yield html[start_idx:end_idx]
                except Exception as e:
                    print(f"Error rendering record {r.id} in bulk print:", e)
                    pass
            
            # Bước 3: Đóng trang với các thẻ script còn lại của trang gốc
            yield base_html[insert_idx:]
            
        # Trả về Response theo dạng stream để tránh Over-Memory Limit và Timeout
        return Response(stream_with_context(generate_bulk_stream()), mimetype="text/html", headers={"Content-Type": "text/html; charset=utf-8"})
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500

# ─── TRÚNG TUYỂN: Toggle trạng thái và gửi sang B ───────────────────
@app.route('/api/history/<int:record_id>/toggle-selected', methods=['POST'])
@auth_required
def api_toggle_selected(record_id):
    try:
        record = FormHistory.query.get(record_id)
        if not record:
            return jsonify({'success': False, 'error': 'Not found'}), 404
        
        data = request.get_json() or {}
        new_donhang = str(data.get('don_hang', '')).strip()
        
        # Nếu field is_selected chưa tồn tại, set default = False
        if not hasattr(record, 'is_selected') or record.is_selected is None:
            record.is_selected = False
        
        # Toggle trạng thái
        new_state = not record.is_selected
        
        if new_state and new_donhang:
            current_jobs_str = (record.don_hang or "").strip()
            final_donhang = new_donhang
            
            if current_jobs_str:
                current_jobs = [j.strip() for j in re.split(r'[,;]+', current_jobs_str) if j.strip()]
                target_upper = unicodedata.normalize('NFC', new_donhang.upper())
                already_has = any(unicodedata.normalize('NFC', j.upper()) == target_upper for j in current_jobs)
                
                if not already_has:
                    current_jobs.append(new_donhang)
                    final_donhang = ", ".join(current_jobs)
                else:
                    final_donhang = current_jobs_str
            
            record.don_hang = final_donhang
            record.is_selected = True
            
        elif not new_state and new_donhang:
            current_jobs_str = (record.don_hang or "").strip()
            if current_jobs_str:
                current_jobs = [j.strip() for j in re.split(r'[,;]+', current_jobs_str) if j.strip()]
                target_upper = unicodedata.normalize('NFC', new_donhang.upper())
                
                new_jobs = []
                for j in current_jobs:
                    if unicodedata.normalize('NFC', j.upper()) != target_upper:
                        new_jobs.append(j)
                        
                final_donhang = ", ".join(new_jobs)
                record.don_hang = final_donhang
                
                if new_jobs:
                    record.is_selected = True # Vẫn trúng tuyển đơn khác
                else:
                    record.is_selected = False
            else:
                record.is_selected = False
                
        elif not new_state and not new_donhang:
            record.is_selected = False

        try:
            if record.data_json:
                jd = json.loads(record.data_json)
                jd['Donhang'] = record.don_hang
                record.data_json = json.dumps(jd, ensure_ascii=False)
        except:
            pass

        db.session.commit()
        
        # Nếu trúng tuyển (is_selected = True), gửi dữ liệu sang B
        if record.is_selected and new_state: # Chỉ gửi khi MỚI bật trúng tuyển
            try:
                form_data = json.loads(record.data_json) if record.data_json else {}
                
                # Mapping dữ liệu A → B
                worker_data = {
                    'id': record.ma_so or f'worker_{record.id}',
                    'full_name': record.ho_ten or '',
                    'date_of_birth': form_data.get('Ngaysinh') or None,
                    'phone_number': form_data.get('Lienhe') or None,
                    'hometown': form_data.get('Noio') or None,
                    'avatar_url': form_data.get('photo') or '',  # Base64 ảnh
                    'win_date': datetime.utcnow().strftime('%Y-%m-%d'),
                    'status': 'DRAFT',
                    'is_placed': False,
                    'passport_expiry': '',
                    'id_card_expiry': '',
                    'health_check_expiry': '',
                    'judicial_record_2_expiry': '',
                }
                
                # Gửi sang B (Firebase)
                b_api_url = os.environ.get('B_API_URL', 'http://localhost:3000')
                response = requests.post(
                    f'{b_api_url}/api/workers/sync-from-a',
                    json=worker_data,
                    timeout=10
                )
                
                if response.status_code != 200:
                    # Nếu gửi thất bại, rollback toggle
                    record.is_selected = False
                    db.session.commit()
                    return jsonify({
                        'success': False,
                        'error': f'Failed to sync to B: {response.text}'
                    }), 500
            except Exception as e:
                # Nếu có lỗi, rollback toggle
                record.is_selected = False
                db.session.commit()
                traceback.print_exc()
                return jsonify({
                    'success': False,
                    'error': f'Sync error: {str(e)}'
                }), 500
        
        return jsonify({
            'success': True,
            'is_selected': record.is_selected,
            'don_hang': record.don_hang,
            'message': 'Đã trúng tuyển' if record.is_selected else 'Đã bỏ trúng tuyển'
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=app.debug, use_reloader=False)